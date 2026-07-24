import csv, math, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(_BASE, "aa_providers_dedup.csv")
OUT = os.path.join(_BASE, "aa_providers_scored.xlsx")

def f(v):
    v = (v or "").strip()
    if v in ("", "None", "null"):
        return None
    try:
        return float(v)
    except Exception:
        return None

rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))
print("loaded rows:", len(rows))

# ---- hierarchical weights (category -> sub-weights) ----
CATS = [
    ("智能体 Agentic", 0.20, [("GDPval-AA", 1.00)]),
    ("编程 Coding",    0.20, [("Terminal-Bench Hard", 0.50),
                              ("Terminal-Bench v2.1", 0.30),
                              ("SciCode", 0.20)]),
    ("通用 General",   0.40, [("LCR", 0.30),
                              ("Omniscience Index", 0.30),
                              ("IFBench", 0.40)]),
    ("知识 Knowledge",  0.20, [("GPQA Diamond", 0.40),
                              ("HLE", 0.60)]),
]
METRICS = [m for _, _, subs in CATS for m, _ in subs]
GLOBAL = {}
for _, cw, subs in CATS:
    for m, sw in subs:
        GLOBAL[m] = cw * sw
assert abs(sum(GLOBAL.values()) - 1.0) < 1e-9, "weights must sum to 1"
print("global weights:", {m: round(GLOBAL[m], 3) for m in METRICS})

# ---- per-metric stats: min/max (min-max norm) + P90/P95 caps + top50 mean ----
stats = {}
P95 = {}
for m in METRICS:
    vals = [f(r.get(m)) for r in rows]
    vals = [v for v in vals if v is not None]
    lo, hi = min(vals), max(vals)
    sv = sorted(vals)
    p90 = sv[min(len(sv) - 1, int(0.90 * (len(sv) - 1)))]
    p95 = sv[min(len(sv) - 1, int(0.95 * (len(sv) - 1)))]
    top50 = sorted(vals)[len(vals) // 2:]
    top50mean = sum(top50) / len(top50)
    stats[m] = (lo, hi, top50mean, p90, p95)
    P95[m] = p95
    print(f"  {m:20} min={lo:.3f} max={hi:.3f} P90={p90:.3f} P95={p95:.3f} top50mean={top50mean:.3f} n={len(vals)}")

# ---- multivariate ridge-regression imputation (iterative) ----
# Predict each metric from ALL OTHER metrics + Intelligence Index (most-complete signal).
# Multiply-missing rows handled by iteration; prediction clipped to P95 (not max).
import numpy as np
II = [f(r.get("Intelligence Index")) for r in rows]
raw = {m: [f(r.get(m)) for r in rows] for m in METRICS}
cur = {m: [(v if v is not None else stats[m][2]) for v in raw[m]] for m in METRICS}  # init w/ top50 mean

ALPHA = 1.0
def feat(i, target):
    xs = [cur[mm][i] for mm in METRICS if mm != target]
    xs.append(0.0 if II[i] is None else II[i])
    return xs

for it in range(30):
    for m in METRICS:
        Xtr, ytr = [], []
        for i in range(len(rows)):
            if raw[m][i] is not None:
                Xtr.append(feat(i, m)); ytr.append(raw[m][i])
        Xtr = np.array(Xtr, dtype=float); ytr = np.array(ytr, dtype=float)
        X1 = np.hstack([np.ones((len(Xtr), 1)), Xtr])
        A = X1.T @ X1 + ALPHA * np.eye(X1.shape[1])
        try:
            beta = np.linalg.solve(A, X1.T @ ytr)
        except np.linalg.LinAlgError:
            beta = np.linalg.lstsq(X1, ytr, rcond=None)[0]
        lo, hi, _, _, p95 = stats[m]
        for i in range(len(rows)):
            if raw[m][i] is None:
                xi = np.array([1.0] + feat(i, m))
                pred = float(xi @ beta)
                cur[m][i] = max(lo, min(p95, pred))   # clip to P95 (avoid inflating to historical best)

def norm(m, v):
    lo, hi, _, _, _ = stats[m]
    return 0.0 if hi == lo else (v - lo) / (hi - lo) * 100.0

# ---- cost assumption ----
INPUT_SHARE, OUTPUT_SHARE = 0.70, 0.30
CACHE_SHARE = 0.50
# effective per-1M-token shares: input(no-cache)=0.35, cache=0.35, output=0.30

out = []
for i, r in enumerate(rows):
    eff = {}; imp_method = {}; imputed = []
    for m in METRICS:
        if raw[m][i] is not None:
            eff[m] = raw[m][i]; imp_method[m] = None
        else:
            eff[m] = cur[m][i]; imputed.append(m); imp_method[m] = "reg"
    nrm = {m: norm(m, eff[m]) for m in METRICS}
    total = sum(GLOBAL[m] * nrm[m] for m in METRICS)

    pin = f(r.get("Price 1M Input"))
    pout = f(r.get("Price 1M Output"))
    pcache = f(r.get("Cache Hit Price"))
    pcache_eff = pcache if pcache is not None else pin
    if None in (pin, pout):
        cost_in = cost_cache = cost_out = cost_total = None
    else:
        cost_in = INPUT_SHARE * (1 - CACHE_SHARE) * pin
        cost_cache = INPUT_SHARE * CACHE_SHARE * pcache_eff
        cost_out = OUTPUT_SHARE * pout
        cost_total = cost_in + cost_cache + cost_out

    out.append({
        "Model": r.get("Model"), "Creator": r.get("Creator"),
        "Reasoning": r.get("Reasoning Model"),
        "Orig Intelligence Index": f(r.get("Intelligence Index")),
        **{m: (round(eff[m], 3) if m in imputed else raw[m][i]) for m in METRICS},
        **{m + " (norm)": round(nrm[m], 1) for m in METRICS},
        "Weighted Total": round(total, 1),
        "Price 1M In": pin, "Price 1M Out": pout, "Cache Hit": pcache_eff,
        "Cost-In $": (round(cost_in, 3) if cost_in is not None else None),
        "Cost-Cache $": (round(cost_cache, 3) if cost_cache is not None else None),
        "Cost-Out $": (round(cost_out, 3) if cost_out is not None else None),
        "Total $/1M": (round(cost_total, 3) if cost_total is not None else None),
        "Imputed": ", ".join(f"{m}({imp_method[m]})" for m in imputed) if imputed else "",
    })

out.sort(key=lambda x: (x["Weighted Total"] if x["Weighted Total"] is not None else -1), reverse=True)

keep = math.ceil(len(out) * 0.15)
out = out[:keep]
print(f"top15% keep = {keep}")

# ---- write xlsx (improved 2-row banded layout) ----
wb = Workbook(); ws = wb.active; ws.title = "Scored"

CAT_COLORS = {
    "智能体 Agentic": ("2E75B6", "DDEBF7"),   # (band dark, cell light tint)
    "编程 Coding":    ("548235", "E2EFDA"),
    "通用 General":   ("BF8F00", "FFF2CC"),
    "知识 Knowledge": ("7030A0", "EAD1DC"),
}
NEUTRAL = "1F4E78"

# group-ordered headers: front = rank+result, base info, metrics by category, cost detail
GROUPS = [
    ("排名与结果", ["Rank", "Model", "Weighted Total", "Total $/1M"]),
    ("基础信息",   ["Creator", "Reasoning", "Orig Intelligence Index"]),
]
for catname, _, subs in CATS:
    GROUPS.append((catname, [m for m, _ in subs] + [m + " (norm)" for m, _ in subs]))
GROUPS.append(("成本明细", ["Price 1M In", "Price 1M Out", "Cache Hit",
                            "Cost-In $", "Cost-Cache $", "Cost-Out $", "Imputed"]))
headers = []
for _, cols in GROUPS:
    headers += cols

# assign rank (1..N by final order)
for idx, row in enumerate(out, 1):
    row["Rank"] = idx

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# row 1: group bands
groups = []
col = 1
for label, cols in GROUPS:
    s = col; e = col + len(cols) - 1
    fill = NEUTRAL
    for catname, _, subs in CATS:
        if label == catname:
            fill = CAT_COLORS[catname][0]
    groups.append((label, s, e, fill))
    col = e + 1
for label, s, e, fill in groups:
    ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
    c = ws.cell(row=1, column=s)
    c.value = label
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

# row 2: field names, tinted by category
metric_light = {}
for ci, h in enumerate(headers, 1):
    for catname, _, subs in CATS:
        if h in ([m for m, _ in subs] + [m + " (norm)" for m, _ in subs]):
            metric_light[ci] = CAT_COLORS[catname][1]
            break
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=c)
    cell.value = h
    fill = metric_light.get(c, NEUTRAL)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color="FFFFFF" if fill == NEUTRAL else "1F4E78", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# data rows
for row in out:
    ws.append([row.get(h) for h in headers])

# styles: zebra + red supplemented cells + total conditional fill
imp_font = Font(color="9C0006", bold=True)
imp_fill = PatternFill("solid", fgColor="FFC7CE")
metric_cols = {m: headers.index(m) + 1 for m in METRICS}
norm_cols = {m: headers.index(m + " (norm)") + 1 for m in METRICS}
zebra = PatternFill("solid", fgColor="F2F2F2")
for ri in range(3, len(out) + 3):
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=ri, column=ci)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ri % 2 == 1:
            cell.fill = zebra
    imp_str = ws.cell(row=ri, column=headers.index("Imputed") + 1).value
    imp_set = set(x.strip() for x in (imp_str or "").split(",") if x.strip())
    for token in imp_set:
        m = token.split("(")[0].strip()
        if m in metric_cols:
            ws.cell(row=ri, column=metric_cols[m]).font = imp_font
            ws.cell(row=ri, column=metric_cols[m]).fill = imp_fill
            ws.cell(row=ri, column=norm_cols[m]).font = imp_font
            ws.cell(row=ri, column=norm_cols[m]).fill = imp_fill
    wt = ws.cell(row=ri, column=headers.index("Weighted Total") + 1).value
    if wt is not None and wt >= 70:
        ws.cell(row=ri, column=headers.index("Weighted Total") + 1).fill = PatternFill("solid", fgColor="C6EFCE")
    elif wt is not None and wt >= 50:
        ws.cell(row=ri, column=headers.index("Weighted Total") + 1).fill = PatternFill("solid", fgColor="FFEB9C")

# number formats
for c, h in enumerate(headers, 1):
    for ri in range(3, len(out) + 3):
        cell = ws.cell(row=ri, column=c)
        if h == "Orig Intelligence Index":
            cell.number_format = "0.0"
        elif h in METRICS:
            cell.number_format = "0.000"
        elif "(norm)" in h:
            cell.number_format = "0.0"
        elif h == "Weighted Total":
            cell.number_format = "0.0"
        elif h in ("Price 1M In", "Price 1M Out", "Cache Hit"):
            cell.number_format = "0.00"
        elif "Cost" in h or h == "Total $/1M":
            cell.number_format = "0.000"

# column widths
widths = {"Rank": 6, "Model": 30, "Creator": 18, "Reasoning": 10,
          "Orig Intelligence Index": 12, "Imputed": 30, "Weighted Total": 13}
for i, h in enumerate(headers, 1):
    L = get_column_letter(i)
    if h in widths:
        ws.column_dimensions[L].width = widths[h]
    elif h in METRICS:
        ws.column_dimensions[L].width = 13
    elif "(norm)" in h:
        ws.column_dimensions[L].width = 10
    elif "Cost" in h or "Price" in h or "Cache" in h or h == "Total $/1M":
        ws.column_dimensions[L].width = 11
    else:
        ws.column_dimensions[L].width = 14
ws.freeze_panes = "E3"
ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 42

# ---- 说明 sheet ----
ns = wb.create_sheet("说明")
title_font = Font(bold=True, size=13, color="1F4E78")
sec_font = Font(bold=True, color="1F4E78")
wrap = Alignment(wrap_text=True, vertical="top")

def line(a, b=""):
    r = ns.max_row + 1
    ns.cell(row=r, column=1, value=a).font = Font(bold=True)
    ns.cell(row=r, column=1).alignment = wrap
    c = ns.cell(row=r, column=2, value=b); c.alignment = wrap

ns["A1"] = "分层加权总分 · 成本估算 — 方法说明"; ns["A1"].font = title_font
line("数据来源", "Artificial Analysis providers leaderboard（2026-07-24 抓取）：1067 模型×服务商 → 按 Model Slug 去重 390 → 取前 15% = 59 行。")
line("总分公式", "Weighted Total = Σ ( 指标归一分 × 全局权重 )，范围 0–100。全局权重 = 大类权重 × 该指标在大类内的子权重。")
line("权重（分层）",
     "智能体 20% → GDPval-AA 100%(=20%)；"
     "编程 20% → Terminal-Bench Hard 50%(10%) / Terminal-Bench v2.1 30%(6%) / SciCode 20%(4%)；"
     "通用 40% → LCR 30%(12%) / Omniscience 30%(12%) / IFBench 40%(16%)；"
     "知识 20% → GPQA Diamond 40%(8%) / HLE 60%(12%)。权重和=1.00。")
line("归一化", "Min-Max：各指标按全量 390 行实测 min/max 线性映射到 0–100。属『相对当前榜单』非理论满分（如 Omniscience 实测区间非 -100~100）。")
line("缺失值预测", "表格无值的指标，用『多变量岭回归 + 迭代填补』预测：以该模型的其他 8 个基准指标 + Intelligence Index 共 9 个特征联合预测目标列（比单用 II 更准，利用指标间交叉相关）；多重缺失的行迭代多轮直至收敛。预测值裁剪到该列 P95（第95百分位），保留梯度同时避免缺失模型被推到历史最高分。红色单元格=被预测值；(norm) 列同步标红；Imputed 列标 (reg)。")
line("成本口径", "每 1M token 消耗：70% 为输入 / 30% 为输出，且 50% 输入命中提示缓存。")
line("成本公式", "Total $/1M = 0.35×输入价 + 0.35×缓存命中价 + 0.30×输出价（USD/百万token）。缓存命中价缺失时回退输入价。价格为各模型被保留变体的服务商报价。")
line("解读提醒", "① 高分代表『在当前样本中靠前』，非理论能力满分。② 含红色格的行其缺失项由回归预测，强模型预测值偏高属预期，参考时仍可优先看白底行。③ 价格为快照，随服务商调价变动。")
# multivariate training R^2 (using final imputed features)
R2 = {}
for m in METRICS:
    Xtr, ytr = [], []
    for i in range(len(rows)):
        if raw[m][i] is not None:
            Xtr.append(feat(i, m)); ytr.append(raw[m][i])
    Xtr = np.array(Xtr, dtype=float); ytr = np.array(ytr, dtype=float)
    X1 = np.hstack([np.ones((len(Xtr), 1)), Xtr])
    beta = np.linalg.lstsq(X1, ytr, rcond=None)[0]
    pred = X1 @ beta
    ybar = ytr.mean()
    ss_tot = ((ytr - ybar) ** 2).sum() or 1e-12
    ss_res = ((ytr - pred) ** 2).sum()
    R2[m] = max(0.0, 1 - ss_res / ss_tot)
r2txt = "；".join(f"{m}={R2[m]:.2f}" for m in METRICS)
line("模型拟合质量 R²", "各指标由『其他 8 指标 + Intelligence Index』多变量岭回归的训练集 R²（越高预测越可靠）：" + r2txt + "。R² 偏低者其预测值仅供参考。")
ns.column_dimensions["A"].width = 16; ns.column_dimensions["B"].width = 110

wb.save(OUT)
print("Saved", OUT)
