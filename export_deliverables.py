import csv, os
from openpyxl import load_workbook

_BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(_BASE, "aa_providers_scored.xlsx")
CSV  = os.path.join(_BASE, "aa_providers_scored.csv")

ws = load_workbook(XLSX)["Scored"]
headers = [c.value for c in ws[2]]
rows = [[ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        for r in range(3, ws.max_row + 1)]

# export full CSV (GitHub renders CSV inline as a table)
with open(CSV, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(headers)
    w.writerows(rows)
print("CSV rows:", len(rows), "cols:", len(headers))

# Top-15 preview for README
def g(name):
    return headers.index(name) + 1
print("\nTop15:")
for r in rows[:15]:
    rank = r[g("Rank") - 1]
    model = r[g("Model") - 1]
    creator = r[g("Creator") - 1]
    total = r[g("Weighted Total") - 1]
    cost = r[g("Total $/1M") - 1]
    imp = r[g("Imputed") - 1] or ""
    print(f'{rank}|{model}|{creator}|{total}|{cost}|{imp}')
