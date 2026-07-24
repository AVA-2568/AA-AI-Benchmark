import csv, math
import numpy as np
from openpyxl import load_workbook

SRC = r"E:\WK\2026-07-24-11-00-42\aa_providers_dedup.csv"
XLSX = r"E:\WK\2026-07-24-11-00-42\aa_providers_scored.xlsx"

def f(v):
    v = (v or "").strip()
    if v in ("", "None", "null"):
        return None
    try:
        return float(v)
    except Exception:
        return None

rows = list(csv.DictReader(open(SRC, encoding="utf-8-sig")))

CATS = [
    ("智能体", 0.20, [("GDPval-AA", 1.00)]),
    ("编程", 0.20, [("Terminal-Bench Hard", 0.50), ("Terminal-Bench v2.1", 0.30), ("SciCode", 0.20)]),
    ("通用", 0.40, [("LCR", 0.30), ("Omniscience Index", 0.30), ("IFBench", 0.40)]),
    ("知识", 0.20, [("GPQA Diamond", 0.40), ("HLE", 0.60)]),
]
METRICS = [m for _, _, subs in CATS for m, _ in subs]
GLOBAL = {}
for _, cw, subs in CATS:
    for m, sw in subs:
        GLOBAL[m] = cw * sw

# stats + P95
stats = {}; P95 = {}
for m in METRICS:
    vals = [f(r.get(m)) for r in rows]; vals = [v for v in vals if v is not None]
    lo, hi = min(vals), max(vals)
    sv = sorted(vals); p95 = sv[min(len(sv)-1, int(0.95*(len(sv)-1)))]
    top50 = sv[len(sv)//2:]; top50mean = sum(top50)/len(top50)
    stats[m] = (lo, hi, top50mean, p95)
    P95[m] = p95

# multivariate ridge iterative imputation (mirror score_aa)
II = [f(r.get("Intelligence Index")) for r in rows]
raw = {m: [f(r.get(m)) for r in rows] for m in METRICS}
cur = {m: [(v if v is not None else stats[m][2]) for v in raw[m]] for m in METRICS}
ALPHA = 1.0
def feat(i, target):
    xs = [cur[mm][i] for mm in METRICS if mm != target]
    xs.append(0.0 if II[i] is None else II[i])
    return xs
for it in range(30):
    for m in METRICS:
        Xtr, ytr = [], []
        for i in range(len(rows)):
            if raw[m][i] is not None:
                Xtr.append(feat(i, m)); ytr.append(raw[m][i])
        Xtr = np.array(Xtr, dtype=float); ytr = np.array(ytr, dtype=float)
        X1 = np.hstack([np.ones((len(Xtr), 1)), Xtr])
        A = X1.T @ X1 + ALPHA * np.eye(X1.shape[1])
        try:
            beta = np.linalg.solve(A, X1.T @ ytr)
        except np.linalg.LinAlgError:
            beta = np.linalg.lstsq(X1, ytr, rcond=None)[0]
        lo, hi, _, p95 = stats[m]
        for i in range(len(rows)):
            if raw[m][i] is None:
                xi = np.array([1.0] + feat(i, m))
                pred = float(xi @ beta)
                cur[m][i] = max(lo, min(p95, pred))

# compute totals
def norm(m, v):
    lo, hi, _, _ = stats[m]
    return 0.0 if hi == lo else (v - lo) / (hi - lo) * 100.0
res = []
for i, r in enumerate(rows):
    nrm = {m: norm(m, cur[m][i]) for m in METRICS}
    total = sum(GLOBAL[m] * nrm[m] for m in METRICS)
    res.append((r["Model"], round(total, 1)))
res.sort(key=lambda x: x[1], reverse=True)
kept = res[:math.ceil(len(res)*0.15)]

# compare with xlsx
ws = load_workbook(XLSX)["Scored"]; hdr = [c.value for c in ws[2]]
xlsx = {}
for rr in range(3, ws.max_row+1):
    nm = ws.cell(row=rr, column=hdr.index("Model")+1).value
    xlsx[nm] = (ws.cell(row=rr, column=hdr.index("Weighted Total")+1).value,
                [ws.cell(row=rr, column=hdr.index(m)+1).value for m in METRICS])

mism = 0
for nm, tot in kept:
    if nm not in xlsx:
        print("  MISSING model in xlsx:", nm); mism += 1; continue
    xtot, xvals = xlsx[nm]
    if abs((xtot or 0) - tot) > 0.05:
        mism += 1
        print(f"  TOTAL MISMATCH {nm}: xlsx={xtot} recompute={tot}")
print(f"kept={len(kept)} | xlsx rows={len(xlsx)} | total mismatches={mism}")
print("Top5 (recompute):", kept[:5])
